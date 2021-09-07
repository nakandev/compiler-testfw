import copy
import openpyxl
import os
import string


class XlsxWriter():
    def __init__(self, config, report, reportdir):
        self.config = config
        self.report = report
        self.reportdir = reportdir
        self.book = openpyxl.Workbook()

    def write(self):
        self.load_template()
        self.write_cover()
        self.write_envinfo()
        self.write_results()
        if hasattr(self.config, 'reportfile'):
            fname = self.config.reportfile + '.xlsx'
            report_path = os.path.join(self.reportdir, fname)
        else:
            report_path = os.path.join(self.reportdir, 'report.xlsx')
        self.book.save(report_path)

    def load_template(self):
        self.book = openpyxl.load_workbook(self.config.report_template_xlsx)

    def write_cover(self):
        if 'cover' not in self.book.sheetnames:
            self.book.create_sheet(title='cover')
        sheet = self.book['cover']
        for row in sheet.rows:
            for cell in row:
                tstr = cell.value
                if tstr is None:
                    continue
                try:
                    cell.value = str(tstr).format(report=self.report)
                except Exception:
                    pass

    def write_envinfo(self):
        if 'env' not in self.book.sheetnames:
            self.book.create_sheet(title='env')
        sheet = self.book['env']
        envinfo = self.report.envinfo
        i = 1
        for info_key, info in envinfo.items():
            sheet.cell(i, 1).value = '[%s]' % (info_key)
            i += 1
            for key, val in info.items():
                sheet.cell(i, 1).value = key
                if isinstance(val, dict):
                    for vk, vv in val.items():
                        sheet.cell(i, 2).value = str(vk)
                        sheet.cell(i, 3).value = str(vv)
                        i += 1
                elif isinstance(val, list):
                    for vv in val:
                        sheet.cell(i, 2).value = str(vv)
                        i += 1
                else:
                    sheet.cell(i, 2).value = str(val)
                    i += 1

    def write_results(self):
        if 'result' not in self.book.sheetnames:
            self.book.create_sheet(title='result')
        sheet = self.book['result']
        tc = self.report.testcases[0]
        row = 0
        maxcol = self._expand_result_row(sheet, row, tc)
        for row, tc in enumerate(self.report.testcases):
            for col in range(sheet.max_column):
                copyfrom = sheet.cell(2, col + 1)
                copyto = sheet.cell(row + 2, col + 1)
                copyto.value = copyfrom.value
        for i, tc in enumerate(self.report.testcases):
            self._expand_result_row(sheet, i + 1, tc, maxcol=maxcol)

    def _str_format(self, fmt, **kwargs):
        formatter = string.Formatter()
        text = ''
        for ret_p in formatter.parse(fmt):
            literal, field, fmtspec, conversion = ret_p
            text += literal
            if field is not None:
                ret_f = formatter.get_field(field, [], kwargs)
                obj, used_key = ret_f
                if isinstance(obj, list):
                    return obj
                else:
                    field_text = str(obj)
                text += field_text
        return text

    def _expand_result_row(self, sheet, row, testcase, maxcol=None):
        max_column = sheet.max_column if maxcol is None else maxcol
        col = max_column - 1
        while col >= 0:
            cell = sheet.cell(row + 1, col + 1)
            if cell.value:
                text = self._str_format(cell.value, testcase=testcase)
                if isinstance(text, list):
                    coodinate = '{}{}:{}{}'.format(
                        openpyxl.utils.get_column_letter(col + 2),
                        row + 1,
                        openpyxl.utils.get_column_letter(max_column + 1),
                        row + 1,
                    )
                    sheet.move_range(coodinate, rows=0, cols=len(text) - 1)
                    if maxcol is None:
                        max_column += len(text)
                    for t in range(len(text)):
                        copyfrom = sheet.cell(row + 1, col + 1)
                        copyto = sheet.cell(row + 1, col + t + 1)
                        copyto.value = str(text[t])
                        if copyfrom.has_style:
                            copyto._style = copy.copy(copyfrom._style)
                else:
                    cell.value = text
            col -= 1
        return max_column
